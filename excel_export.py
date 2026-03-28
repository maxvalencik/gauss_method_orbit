"""
Excel export — detailed step-by-step workbook for Gauss's method.
Sheets:
  1. Summary          — inputs, final state, orbital elements
  2. Algorithm 5.5    — every intermediate quantity
  3. Algorithm 5.6    — per-iteration refinement table  (if n_iter > 0)
  4. Orbital Elements — COEs with derivation quantities
"""

import numpy as np
from math import radians, sqrt, cos, sin, degrees, asin, atan2, pi, acos
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

from gauss_method import (
    _unit_vector, _lagrange_fg, _compute_D, MU, RE,
    state_to_elements,
)
from app import (
    _C, _S, _universal_kepler, _fg,
    observer_R_from_lat_lst,
)

# ─────────────────────────────────────────────────────────────────────────────
# Colour palette  (dark-themed matching the web UI)
# ─────────────────────────────────────────────────────────────────────────────
C_TITLE   = "1F2937"   # dark header
C_HEAD    = "1E3A5F"   # section header
C_SUBHD   = "0D2137"   # step header
C_ALT     = "0F1923"   # alternating row
C_BG      = "0D1117"   # default bg
C_TEXT    = "E6EDF3"
C_ACCENT  = "F0A500"   # gold
C_GREEN   = "56D364"
C_BLUE    = "58A6FF"
C_ORANGE  = "FFA657"

WHITE = "FFFFFF"

def _fill(hex_col):
    return PatternFill("solid", fgColor=hex_col)

def _font(bold=False, color=C_TEXT, size=10):
    return Font(bold=bold, color=color, size=size, name="Calibri")

def _border():
    s = Side(style="thin", color="30363D")
    return Border(left=s, right=s, top=s, bottom=s)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

NUM4  = '#,##0.0000'
NUM6  = '#,##0.000000'
NUM2  = '#,##0.00'
SCI   = '0.000000E+00'


# ─────────────────────────────────────────────────────────────────────────────
# Low-level cell writers
# ─────────────────────────────────────────────────────────────────────────────
def _write(ws, row, col, value, bold=False, color=C_TEXT, bg=None,
           fmt=None, align="left", size=10):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(bold=bold, color=color, size=size, name="Calibri")
    cell.alignment = _center() if align == "center" else _left()
    if bg:
        cell.fill = _fill(bg)
    if fmt:
        cell.number_format = fmt
    cell.border = _border()
    return cell


def _title_row(ws, row, text, ncols=8):
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row,   end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font      = Font(bold=True, color=C_ACCENT, size=13, name="Calibri")
    c.fill      = _fill(C_TITLE)
    c.alignment = _center()
    c.border    = _border()


def _section(ws, row, text, ncols=8, color=C_HEAD):
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row,   end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font      = Font(bold=True, color=WHITE, size=11, name="Calibri")
    c.fill      = _fill(color)
    c.alignment = _left()
    c.border    = _border()


def _step(ws, row, step_num, text, ncols=8):
    _section(ws, row, f"  Step {step_num}:  {text}", ncols=ncols, color=C_SUBHD)


def _kv(ws, row, label, value, unit="", fmt=None, bg=C_BG):
    """Key-value row: col1=label, col2=value, col3=unit."""
    _write(ws, row, 1, label,  bold=True,  bg=bg, align="right")
    _write(ws, row, 2, value,  bold=False, bg=bg, fmt=fmt)
    if unit:
        _write(ws, row, 3, unit, bold=False, color="8B949E", bg=bg)


def _vec_row(ws, row, label, vec, unit="", fmt=NUM4, bg=C_BG):
    """Label | x | y | z | unit"""
    _write(ws, row, 1, label, bold=True,  align="right", bg=bg)
    _write(ws, row, 2, float(vec[0]), fmt=fmt, bg=bg)
    _write(ws, row, 3, float(vec[1]), fmt=fmt, bg=bg)
    _write(ws, row, 4, float(vec[2]), fmt=fmt, bg=bg)
    if unit:
        _write(ws, row, 5, unit, color="8B949E", bg=bg)


def _matrix_row(ws, row, label, mat_row, fmt=NUM4, bg=C_BG):
    _write(ws, row, 1, label, bold=True, align="right", bg=bg)
    for j, v in enumerate(mat_row, start=2):
        _write(ws, row, j, float(v), fmt=fmt, bg=bg)


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 1 — Summary
# ─────────────────────────────────────────────────────────────────────────────
def _sheet_summary(wb, inputs, steps56):
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False

    obs   = inputs["obs"]
    Rs    = inputs["Rs"]
    r     = steps56["r2_final"]
    v     = steps56["v2_final"]
    rhos  = steps56["rhos_final"]
    elems = state_to_elements(r, v)

    r_mag = float(np.linalg.norm(r))
    v_mag = float(np.linalg.norm(v))

    row = 1
    _title_row(ws, row, "Gauss's Method — Preliminary Orbit Determination", ncols=6)
    row += 1
    _write(ws, row, 1, "Curtis, Orbital Mechanics for Engineering Students, 4th ed.",
           color="8B949E", bg=C_BG)
    row += 2

    # ── Observations ────────────────────────────────────────────────────────
    _section(ws, row, "Observations", ncols=6)
    row += 1
    for hdr, col in [("", 1), ("t (s)", 2), ("RA α (°)", 3),
                     ("Dec δ (°)", 4), ("R vector (km)", 5)]:
        _write(ws, row, col, hdr, bold=True, bg=C_HEAD, align="center",
               color=WHITE)
    row += 1
    for i, (ob, R) in enumerate(zip(obs, Rs), 1):
        bg = C_ALT if i % 2 == 0 else C_BG
        _write(ws, row, 1, f"Obs {i}", bold=True, bg=bg)
        _write(ws, row, 2, ob["t"],   fmt=NUM4, bg=bg)
        _write(ws, row, 3, ob["ra_deg"],  fmt=NUM6, bg=bg)
        _write(ws, row, 4, ob["dec_deg"], fmt=NUM6, bg=bg)
        _write(ws, row, 5,
               f"[{R[0]:.3f}, {R[1]:.3f}, {R[2]:.3f}]", bg=bg)
        row += 1

    row += 1

    # ── Algorithm used ───────────────────────────────────────────────────────
    n_iter = inputs.get("n_iter", 0)
    alg = "Algorithm 5.5 only" if n_iter == 0 else \
          f"Algorithm 5.5  +  Algorithm 5.6 ({n_iter} refinement iterations)"
    _section(ws, row, "Algorithm Used", ncols=6)
    row += 1
    _write(ws, row, 1, alg, bg=C_BG)
    row += 2

    # ── State vector ─────────────────────────────────────────────────────────
    _section(ws, row, f"State Vector at t₂ = {obs[1]['t']:.2f} s", ncols=6)
    row += 1
    for hdr, col in [("", 1), ("x", 2), ("y", 3), ("z", 4), ("|·|", 5), ("unit", 6)]:
        _write(ws, row, col, hdr, bold=True, bg=C_HEAD, color=WHITE, align="center")
    row += 1
    _vec_row(ws, row, "r₂  (km)", r, fmt=NUM4)
    _write(ws, row, 5, r_mag, fmt=NUM4, bg=C_BG)
    _write(ws, row, 6, "km",  color="8B949E", bg=C_BG)
    row += 1
    _vec_row(ws, row, "v₂  (km/s)", v, fmt=NUM6, bg=C_ALT)
    _write(ws, row, 5, v_mag, fmt=NUM6, bg=C_ALT)
    _write(ws, row, 6, "km/s", color="8B949E", bg=C_ALT)
    row += 1
    _kv(ws, row, "ρ₁", rhos[0], "km", fmt=NUM4)
    row += 1
    _kv(ws, row, "ρ₂", rhos[1], "km", fmt=NUM4, bg=C_ALT)
    row += 1
    _kv(ws, row, "ρ₃", rhos[2], "km", fmt=NUM4)
    row += 2

    # ── Orbital elements ──────────────────────────────────────────────────────
    _section(ws, row, "Classical Orbital Elements  (Algorithm 4.2)", ncols=6)
    row += 1
    items = [
        ("Semi-major axis",          "a",    elems["a"],    "km",    NUM4),
        ("Eccentricity",             "e",    elems["e"],    "—",     NUM6),
        ("Inclination",              "i",    elems["i"],    "deg",   NUM4),
        ("RAAN",                     "Ω",    elems["raan"], "deg",   NUM4),
        ("Argument of perigee",      "ω",    elems["argp"], "deg",   NUM4),
        ("True anomaly at t₂",       "ν",    elems["nu"],   "deg",   NUM4),
        ("Orbital period",           "T",    elems["T"],    "s",     NUM2),
        ("Orbital period",           "T",    elems["T"]/3600, "hr",  NUM4),
    ]
    for k, (label, sym, val, unit, fmt) in enumerate(items):
        bg = C_ALT if k % 2 else C_BG
        _write(ws, row, 1, label, bold=True, align="right", bg=bg)
        _write(ws, row, 2, sym, bold=True, color=C_ACCENT, align="center", bg=bg)
        _write(ws, row, 3, val, fmt=fmt, bg=bg)
        _write(ws, row, 4, unit, color="8B949E", bg=bg)
        row += 1

    # column widths
    for col, w in [(1,22),(2,16),(3,16),(4,16),(5,16),(6,8)]:
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A3"


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 2 — Algorithm 5.5 step-by-step
# ─────────────────────────────────────────────────────────────────────────────
def _sheet_alg55(wb, inputs, steps55):
    ws = wb.create_sheet("Algorithm 5.5")
    ws.sheet_view.showGridLines = False

    obs  = inputs["obs"]
    Rs   = inputs["Rs"]
    N    = 8   # number of columns to use

    row = 1
    _title_row(ws, row, "Algorithm 5.5 — Gauss's Method  (Step by Step)", ncols=N)
    row += 2

    # ── Step 1: time intervals ────────────────────────────────────────────────
    _step(ws, row, 1, "Compute time intervals  τ₁ = t₁ − t₂,  τ₃ = t₃ − t₂", ncols=N)
    row += 1
    tau1, tau3 = steps55["tau1"], steps55["tau3"]
    _kv(ws, row, "τ₁ = t₁ − t₂", tau1, "s", fmt=NUM4)
    row += 1
    _kv(ws, row, "τ₃ = t₃ − t₂", tau3, "s", fmt=NUM4, bg=C_ALT)
    row += 2

    # ── Step 2: unit vectors ──────────────────────────────────────────────────
    _step(ws, row, 2, "Direction-cosine unit vectors  L̂ᵢ = [cos δᵢ cos αᵢ,  cos δᵢ sin αᵢ,  sin δᵢ]", ncols=N)
    row += 1
    for col, hdr in [(1,""), (2,"l  (x)"), (3,"m  (y)"), (4,"n  (z)")]:
        _write(ws, row, col, hdr, bold=True, bg=C_HEAD, color=WHITE, align="center")
    row += 1
    for k, (label, L) in enumerate([("L̂₁", steps55["L1"]),
                                      ("L̂₂", steps55["L2"]),
                                      ("L̂₃", steps55["L3"])]):
        bg = C_ALT if k % 2 else C_BG
        _vec_row(ws, row, label, L, fmt=NUM6, bg=bg)
        row += 1
    row += 1

    # ── Step 3: cross products ────────────────────────────────────────────────
    _step(ws, row, 3, "Cross products  p₁ = L̂₂×L̂₃,  p₂ = L̂₁×L̂₃,  p₃ = L̂₁×L̂₂", ncols=N)
    row += 1
    for col, hdr in [(1,""), (2,"x"), (3,"y"), (4,"z")]:
        _write(ws, row, col, hdr, bold=True, bg=C_HEAD, color=WHITE, align="center")
    row += 1
    for k, (label, p) in enumerate([("p₁", steps55["p1"]),
                                      ("p₂", steps55["p2"]),
                                      ("p₃", steps55["p3"])]):
        bg = C_ALT if k % 2 else C_BG
        _vec_row(ws, row, label, p, fmt=NUM6, bg=bg)
        row += 1
    row += 1

    # ── Step 4: D0 ────────────────────────────────────────────────────────────
    _step(ws, row, 4, "Scalar triple product  D₀ = L̂₁ · p₁", ncols=N)
    row += 1
    _kv(ws, row, "D₀", steps55["D0"], "", fmt=NUM6)
    row += 2

    # ── Step 5: D matrix ──────────────────────────────────────────────────────
    _step(ws, row, 5, "D matrix  Dᵢⱼ = Rⱼ · pᵢ", ncols=N)
    row += 1
    D = steps55["D"]
    for col, hdr in [(1,""), (2,"R₁ · pᵢ"), (3,"R₂ · pᵢ"), (4,"R₃ · pᵢ")]:
        _write(ws, row, col, hdr, bold=True, bg=C_HEAD, color=WHITE, align="center")
    row += 1
    for k in range(3):
        bg = C_ALT if k % 2 else C_BG
        _matrix_row(ws, row, f"  p{k+1}  (row {k+1})", D[k], fmt=NUM4, bg=bg)
        row += 1
    row += 1

    # ── Step 6: solve F(r₂)=0 ────────────────────────────────────────────────
    _step(ws, row, 6,
          "Solve  F(r₂) = |ρ₂(r₂) L̂₂ + R₂| − r₂ = 0  (bisection)", ncols=N)
    row += 1
    _write(ws, row, 1,
           "ρ₂(r₂) = (D₂₂ − c₁D₂₁ − c₃D₂₃) / D₀   with   "
           "c₁ = g₃/W,  c₃ = −g₁/W,  W = f₁g₃ − f₃g₁",
           color="8B949E", bg=C_BG)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=N)
    row += 1
    _kv(ws, row, "r₂  (solution)", steps55["r2_sol"], "km", fmt=NUM4)
    row += 2

    # ── Step 7: Lagrange f,g at solution ─────────────────────────────────────
    _step(ws, row, 7,
          "Lagrange f, g  (2nd-order Taylor)  at r₂ solution", ncols=N)
    row += 1
    for col, hdr in [(1,""), (2,"f"), (3,"g")]:
        _write(ws, row, col, hdr, bold=True, bg=C_HEAD, color=WHITE, align="center")
    row += 1
    for k, (label, f, g) in enumerate([
            ("τ₁  (obs 1)", steps55["f1"], steps55["g1"]),
            ("τ₃  (obs 3)", steps55["f3"], steps55["g3"])]):
        bg = C_ALT if k else C_BG
        _write(ws, row, 1, label, bold=True, align="right", bg=bg)
        _write(ws, row, 2, f, fmt=NUM6, bg=bg)
        _write(ws, row, 3, g, fmt=NUM6, bg=bg)
        row += 1
    _kv(ws, row, "W = f₁g₃ − f₃g₁", steps55["W"], "", fmt=NUM6)
    row += 2

    # ── Step 8: c1, c3 ───────────────────────────────────────────────────────
    _step(ws, row, 8, "Scalar multipliers  c₁ = g₃/W,  c₃ = −g₁/W", ncols=N)
    row += 1
    _kv(ws, row, "c₁", steps55["c1"], "", fmt=NUM6)
    row += 1
    _kv(ws, row, "c₃", steps55["c3"], "", fmt=NUM6, bg=C_ALT)
    row += 2

    # ── Step 9: slant ranges ─────────────────────────────────────────────────
    _step(ws, row, 9, "Slant ranges  ρᵢ", ncols=N)
    row += 1
    for k, (label, val) in enumerate([
            ("ρ₁ = (D₁₂ − c₁D₁₁ − c₃D₁₃) / (c₁D₀)", steps55["rho1"]),
            ("ρ₂ = (D₂₂ − c₁D₂₁ − c₃D₂₃) / D₀",     steps55["rho2"]),
            ("ρ₃ = (D₃₂ − c₁D₃₁ − c₃D₃₃) / (c₃D₀)", steps55["rho3"])]):
        bg = C_ALT if k % 2 else C_BG
        _write(ws, row, 1, label, bold=True, align="right", bg=bg)
        _write(ws, row, 2, val, fmt=NUM4, bg=bg)
        _write(ws, row, 3, "km", color="8B949E", bg=bg)
        row += 1
    row += 1

    # ── Step 10: position vectors ─────────────────────────────────────────────
    _step(ws, row, 10,
          "Position vectors  rᵢ = ρᵢ L̂ᵢ + Rᵢ", ncols=N)
    row += 1
    for col, hdr in [(1,""), (2,"x (km)"), (3,"y (km)"), (4,"z (km)"),
                     (5,"|r| (km)")]:
        _write(ws, row, col, hdr, bold=True, bg=C_HEAD, color=WHITE, align="center")
    row += 1
    for k, (label, rv) in enumerate([("r₁", steps55["r1"]),
                                       ("r₂", steps55["r2"]),
                                       ("r₃", steps55["r3"])]):
        bg = C_ALT if k % 2 else C_BG
        _vec_row(ws, row, label, rv, fmt=NUM4, bg=bg)
        _write(ws, row, 5, float(np.linalg.norm(rv)), fmt=NUM4, bg=bg)
        row += 1
    row += 1

    # ── Step 11: velocity ─────────────────────────────────────────────────────
    _step(ws, row, 11,
          "Velocity at t₂:  v₂ = (−f₃ r₁ + f₁ r₃) / W", ncols=N)
    row += 1
    for col, hdr in [(1,""), (2,"x (km/s)"), (3,"y (km/s)"), (4,"z (km/s)"),
                     (5,"|v₂| (km/s)")]:
        _write(ws, row, col, hdr, bold=True, bg=C_HEAD, color=WHITE, align="center")
    row += 1
    v2 = steps55["v2"]
    _vec_row(ws, row, "v₂", v2, fmt=NUM6)
    _write(ws, row, 5, float(np.linalg.norm(v2)), fmt=NUM6, bg=C_BG)

    # column widths
    for col, w in [(1,32),(2,16),(3,16),(4,16),(5,16),(6,6),(7,6),(8,6)]:
        ws.column_dimensions[get_column_letter(col)].width = w


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 3 — Algorithm 5.6 refinement table
# ─────────────────────────────────────────────────────────────────────────────
def _sheet_alg56(wb, iters):
    if not iters:
        return
    ws = wb.create_sheet("Algorithm 5.6")
    ws.sheet_view.showGridLines = False

    N = 16
    row = 1
    _title_row(ws, row, "Algorithm 5.6 — Iterative Refinement (Universal Variables)", ncols=N)
    row += 2

    headers = [
        "Iter", "r₂ (km)", "v₂ (km/s)", "α (1/km)",
        "v_r (km/s)", "χ₁", "χ₃",
        "f₁", "g₁", "f₃", "g₃",
        "c₁", "c₃",
        "ρ₁ (km)", "ρ₂ (km)", "ρ₃ (km)",
    ]
    for j, h in enumerate(headers, 1):
        _write(ws, row, j, h, bold=True, bg=C_HEAD, color=WHITE, align="center")
    row += 1

    for k, it in enumerate(iters):
        bg = C_ALT if k % 2 else C_BG
        vals = [
            k + 1,
            it["r2_mag"], it["v2_mag"], it["alpha"],
            it["vr"], it["chi1"], it["chi3"],
            it["f1"], it["g1"], it["f3"], it["g3"],
            it["c1"], it["c3"],
            it["rho1"], it["rho2"], it["rho3"],
        ]
        fmts = [None, NUM4, NUM6, SCI, NUM6, NUM4, NUM4,
                NUM6, NUM6, NUM6, NUM6, NUM6, NUM6,
                NUM4, NUM4, NUM4]
        for j, (v, f) in enumerate(zip(vals, fmts), 1):
            _write(ws, row, j, v, fmt=f, bg=bg)
        row += 1

    # column widths
    for col, w in enumerate(
        [6, 14, 14, 14, 12, 14, 14, 12, 12, 12, 12, 14, 14, 12, 12, 12], 1):
        ws.column_dimensions[get_column_letter(col)].width = w


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 4 — Orbital Elements derivation
# ─────────────────────────────────────────────────────────────────────────────
def _sheet_elements(wb, r_vec, v_vec):
    ws = wb.create_sheet("Orbital Elements")
    ws.sheet_view.showGridLines = False

    mu = MU
    r  = np.linalg.norm(r_vec)
    v  = np.linalg.norm(v_vec)
    h_vec = np.cross(r_vec, v_vec)
    h     = np.linalg.norm(h_vec)
    n_vec = np.cross([0., 0., 1.], h_vec)
    n     = np.linalg.norm(n_vec)
    e_vec = ((v**2 - mu/r)*r_vec - np.dot(r_vec, v_vec)*v_vec) / mu
    e     = np.linalg.norm(e_vec)
    eps   = 0.5*v**2 - mu/r
    a     = -mu/(2*eps) if abs(eps) > 1e-12 else float('inf')
    elems = state_to_elements(r_vec, v_vec)

    N = 7
    row = 1
    _title_row(ws, row, "Orbital Elements — Algorithm 4.2  (Derivation)", ncols=N)
    row += 2

    # Intermediate quantities
    _section(ws, row, "Intermediate Quantities", ncols=N)
    row += 1
    items = [
        ("|r₂|",       r,                         "km",      NUM4),
        ("|v₂|",       v,                         "km/s",    NUM6),
        ("h_vec = r₂×v₂", None,                  "km²/s",   None),
        ("  hx",       h_vec[0],                  "km²/s",   NUM4),
        ("  hy",       h_vec[1],                  "km²/s",   NUM4),
        ("  hz",       h_vec[2],                  "km²/s",   NUM4),
        ("|h|",        h,                          "km²/s",   NUM4),
        ("n_vec = K̂×h_vec", None,               "km²/s",   None),
        ("  nx",       n_vec[0],                  "",        NUM4),
        ("  ny",       n_vec[1],                  "",        NUM4),
        ("  nz",       n_vec[2],                  "",        NUM4),
        ("|n|",        n,                          "",        NUM4),
        ("e_vec (eccentricity vector)", None,      "",        None),
        ("  ex",       e_vec[0],                  "",        NUM6),
        ("  ey",       e_vec[1],                  "",        NUM6),
        ("  ez",       e_vec[2],                  "",        NUM6),
        ("|e|",        e,                          "—",       NUM6),
        ("ε (specific energy)", eps,               "km²/s²",  NUM6),
    ]
    for k, (label, val, unit, fmt) in enumerate(items):
        bg = C_ALT if k % 2 else C_BG
        _write(ws, row, 1, label, bold=True, align="right", bg=bg)
        if val is not None:
            _write(ws, row, 2, float(val), fmt=fmt, bg=bg)
        _write(ws, row, 3, unit, color="8B949E", bg=bg)
        row += 1
    row += 1

    # Elements
    _section(ws, row, "Classical Orbital Elements", ncols=N)
    row += 1
    coe = [
        ("a  — semi-major axis",     elems["a"],    "km",  NUM4,
         "a = −μ / (2ε)"),
        ("e  — eccentricity",        elems["e"],    "—",   NUM6,
         "e = |e_vec|"),
        ("i  — inclination",         elems["i"],    "deg", NUM4,
         "cos i = hz / |h|"),
        ("Ω  — RAAN",                elems["raan"], "deg", NUM4,
         "cos Ω = nx / |n|  (if ny<0 → 360−Ω)"),
        ("ω  — arg of perigee",      elems["argp"], "deg", NUM4,
         "cos ω = (n·e)/(|n||e|)  (if ez<0 → 360−ω)"),
        ("ν  — true anomaly",        elems["nu"],   "deg", NUM4,
         "cos ν = (e·r)/(|e||r|)  (if r·v<0 → 360−ν)"),
        ("T  — orbital period",      elems["T"],    "s",   NUM2,
         "T = 2π√(a³/μ)"),
    ]
    for col, hdr in [(1,"Element"), (2,"Value"), (3,"Unit"), (4,"Formula")]:
        _write(ws, row, col, hdr, bold=True, bg=C_HEAD, color=WHITE, align="center")
    row += 1
    for k, (label, val, unit, fmt, formula) in enumerate(coe):
        bg = C_ALT if k % 2 else C_BG
        _write(ws, row, 1, label, bold=True, align="right", bg=bg)
        _write(ws, row, 2, float(val), fmt=fmt, bg=bg)
        _write(ws, row, 3, unit, color="8B949E", bg=bg)
        _write(ws, row, 4, formula, color="8B949E", bg=bg)
        row += 1

    for col, w in [(1,28),(2,16),(3,8),(4,40)]:
        ws.column_dimensions[get_column_letter(col)].width = w


# ─────────────────────────────────────────────────────────────────────────────
# Main entry point
# ─────────────────────────────────────────────────────────────────────────────
def build_workbook(inputs: dict) -> bytes:
    """
    Re-run every algorithm step with full verbosity and produce an Excel
    workbook. Returns raw bytes ready to send as a file download.

    inputs dict keys:
        obs        list of 3 dicts  {t, ra_deg, dec_deg}
        Rs         list of 3 np.ndarray  (observer ECI km)
        n_iter     int  (number of Algorithm 5.6 iterations)
    """
    obs   = inputs["obs"]
    Rs    = inputs["Rs"]
    n_iter = inputs.get("n_iter", 0)

    # Convert to radians
    Ls = []
    for ob in obs:
        ra  = radians(ob["ra_deg"])
        dec = radians(ob["dec_deg"])
        Ls.append(_unit_vector(ra, dec))
    L1, L2, L3 = Ls
    R1, R2, R3 = [np.array(R) for R in Rs]

    t1, t2, t3 = [ob["t"] for ob in obs]
    tau1 = t1 - t2
    tau3 = t3 - t2

    D0, D = _compute_D(L1, L2, L3, R1, R2, R3)
    p1 = np.cross(L2, L3)
    p2 = np.cross(L1, L3)
    p3 = np.cross(L1, L2)

    # ── Bisection to find r2 ────────────────────────────────────────────────
    def rho2_of_r2(r2_try):
        f1, g1 = _lagrange_fg(tau1, r2_try)
        f3, g3 = _lagrange_fg(tau3, r2_try)
        W   = f1*g3 - f3*g1
        c1  = g3 / W
        c3  = -g1 / W
        return (D[1,1] - c1*D[1,0] - c3*D[1,2]) / D0

    def F(r2_try):
        rho2 = rho2_of_r2(r2_try)
        return float(np.linalg.norm(rho2 * L2 + R2)) - r2_try

    r2_grid = np.linspace(RE + 100, 500_000, 2000)
    r2_sol = None
    for k in range(len(r2_grid)-1):
        if F(r2_grid[k]) * F(r2_grid[k+1]) < 0:
            a_b, b_b = r2_grid[k], r2_grid[k+1]
            for _ in range(200):
                m = 0.5*(a_b + b_b)
                if F(a_b)*F(m) < 0: b_b = m
                else: a_b = m
                if abs(b_b - a_b) < 1e-10: break
            r2_sol = 0.5*(a_b + b_b)
            break

    f1_s, g1_s = _lagrange_fg(tau1, r2_sol)
    f3_s, g3_s = _lagrange_fg(tau3, r2_sol)
    W_s  = f1_s*g3_s - f3_s*g1_s
    c1_s = g3_s / W_s
    c3_s = -g1_s / W_s
    rho1_s = (D[0,1] - c1_s*D[0,0] - c3_s*D[0,2]) / (c1_s * D0)
    rho2_s = (D[1,1] - c1_s*D[1,0] - c3_s*D[1,2]) / D0
    rho3_s = (D[2,1] - c1_s*D[2,0] - c3_s*D[2,2]) / (c3_s * D0)
    r1_vec = rho1_s * L1 + R1
    r2_vec = rho2_s * L2 + R2
    r3_vec = rho3_s * L3 + R3
    v2_vec = (-f3_s * r1_vec + f1_s * r3_vec) / W_s

    steps55 = {
        "tau1": tau1, "tau3": tau3,
        "L1": L1, "L2": L2, "L3": L3,
        "p1": p1, "p2": p2, "p3": p3,
        "D0": D0, "D": D,
        "r2_sol": r2_sol,
        "f1": f1_s, "g1": g1_s, "f3": f3_s, "g3": g3_s, "W": W_s,
        "c1": c1_s, "c3": c3_s,
        "rho1": rho1_s, "rho2": rho2_s, "rho3": rho3_s,
        "r1": r1_vec, "r2": r2_vec, "r3": r3_vec,
        "v2": v2_vec,
    }

    # ── Algorithm 5.6 iterations ────────────────────────────────────────────
    r2_cur, v2_cur = r2_vec.copy(), v2_vec.copy()
    rho1_cur, rho2_cur, rho3_cur = rho1_s, rho2_s, rho3_s
    iters56 = []
    for _ in range(n_iter):
        r2_mag = float(np.linalg.norm(r2_cur))
        v2_mag = float(np.linalg.norm(v2_cur))
        alpha  = 2.0/r2_mag - v2_mag**2/MU
        vr     = float(np.dot(v2_cur, r2_cur) / r2_mag)
        chi1   = float(_universal_kepler(r2_mag, vr, alpha, tau1))
        chi3   = float(_universal_kepler(r2_mag, vr, alpha, tau3))
        f1i, g1i = _fg(chi1, tau1, r2_mag, alpha)
        f3i, g3i = _fg(chi3, tau3, r2_mag, alpha)
        Wi  = f1i*g3i - f3i*g1i
        c1i =  g3i / Wi
        c3i = -g1i / Wi
        p1i = (D[0,1] - c1i*D[0,0] - c3i*D[0,2]) / (c1i * D0)
        p2i = (D[1,1] - c1i*D[1,0] - c3i*D[1,2]) / D0
        p3i = (D[2,1] - c1i*D[2,0] - c3i*D[2,2]) / (c3i * D0)
        r1i = R1 + p1i*L1
        r2i = R2 + p2i*L2
        r3i = R3 + p3i*L3
        v2i = (-f3i*r1i + f1i*r3i) / Wi
        iters56.append({
            "r2_mag": r2_mag, "v2_mag": v2_mag,
            "alpha": alpha, "vr": vr,
            "chi1": chi1, "chi3": chi3,
            "f1": f1i, "g1": g1i, "f3": f3i, "g3": g3i,
            "c1": c1i, "c3": c3i,
            "rho1": p1i, "rho2": p2i, "rho3": p3i,
        })
        r2_cur, v2_cur = r2i, v2i
        rho1_cur, rho2_cur, rho3_cur = p1i, p2i, p3i

    steps56 = {
        "r2_final": r2_cur,
        "v2_final": v2_cur,
        "rhos_final": (rho1_cur, rho2_cur, rho3_cur),
    }

    # ── Build workbook ───────────────────────────────────────────────────────
    wb = Workbook()
    _sheet_summary(wb, inputs, steps56)
    _sheet_alg55(wb, inputs, steps55)
    if iters56:
        _sheet_alg56(wb, iters56)
    _sheet_elements(wb, r2_cur, v2_cur)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
